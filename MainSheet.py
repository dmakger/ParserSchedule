from typing import Union

from openpyxl import Workbook
from openpyxl.worksheet._write_only import WriteOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet

from BorderCoord import BorderCoord as bc
from FontCoord import FontCoord as fc
from AlignmentCoord import AlignmentCoord as ac
from Coord import Coord


class MainSheet:
    NAME_SHEET = "Ведомость усп. и посещ."

    def __init__(self, wb: Workbook, subjects: dict, speciality: str, group: str, month: int, data: dict):
        self.wb = wb
        self.ws: Union[WriteOnlyWorksheet, Worksheet] = None
        self.subjects = subjects
        self.speciality = speciality
        self.group = group
        self.month = self.get_month(month)

        self.start_col_subjects = 3
        self.end_col_subjects = self.start_col_subjects
        self.last_col = 1
        self.last_row = 4

        self.data = data
        print(data['skip'])
        print(data['subjects'])

        self.create_sheet()

    def get_month(self, num: int):
        if num < 1 or num > 12:
            return None
        months = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                  "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
        return months[num - 1]

    def create_sheet(self):
        ws = self.wb.create_sheet(self.NAME_SHEET)
        self.ws = ws

        self.create_students()
        self.create_subjects()
        self.create_statistic()

        self.create_title()
        self.create_logical()
        self.create_result()

    def create_students(self):
        """Создание записи со студентами"""
        print("Создание записи со студентами")
        # Title
        Coord([3, 1], [4, 1], border=bc.get_border_thin(top=bc.MEDIUM, left=bc.MEDIUM), font=fc(bold=True)) \
            .draw(ws=self.ws, title="№ п/п")
        Coord([3, 2], [4, 2], border=bc.get_border_thin(top=bc.MEDIUM), font=fc(bold=True)) \
            .draw(ws=self.ws, title="ФИО")
        self.last_col = 2

        # Main
        students = list(list(self.subjects.values())[0].keys())
        current_row = 5
        for i in range(len(students)):
            Coord([current_row, 1], border=bc.get_border_thin(left=bc.MEDIUM)).draw(ws=self.ws, title=str(i + 1))
            Coord([current_row, 2], border=bc.get_border_thin()) \
                .draw(ws=self.ws, title=students[i], align=ac(horizontal=ac.LEFT), type_size=Coord.CELL_SIZE_TYPE_MAX)
            current_row += 1

    def create_subjects(self):
        """Создание записи с предметами и оценками"""
        print("Создание титульной записи с предметами и оценками")
        text_mini = fc(size=fc.FONT_SIZE_MINI, bold=True)
        current_row = 4
        num_students = len(list(self.subjects.values())[0].keys())

        data_subjects = self.data['subjects']
        letter_start = Coord.get_column_letter(data_subjects['result_start_col'])
        letter_end = Coord.get_column_letter(data_subjects['col_end'])
        range_subjects = f"{letter_start}{data_subjects['last_row']}:{letter_end}{data_subjects['last_row']}"
        # =ОКРУГЛ(ЕСЛИОШИБКА(СРЗНАЧЕСЛИМН('П1-18 (БН)'!$C4:$EB4;'П1-18 (БН)'!$C$22:$EB$22;'Ведомость усп.и посещ.'!D$4); 2); 0)
        for title_subject in self.subjects:
            letter_subject = Coord.get_column_letter(self.end_col_subjects)
            # Оценки по предмету
            for i in range(1, num_students + 1):
                pos_student = data_subjects['result_start_row'] + i - 1
                title = f"=ROUND(" \
                        f"IFERROR(" \
                        f"AVERAGEIFS('{data_subjects['name_sheet']}'!{letter_start}{pos_student}:{letter_end}{pos_student}," \
                        f"'{data_subjects['name_sheet']}'!{range_subjects}," \
                        f"'{self.NAME_SHEET}'!{letter_subject}{}" \
                        f")," \
                        f"2)," \
                        f"0)"
                print(title, self.end_col_subjects)
                Coord([current_row + i, self.end_col_subjects], border=bc.get_border_thin(), title=title) \
                    .draw(ws=self.ws, type_size=Coord.CELL_SIZE_TYPE_MAX)
            # Название предмета
            Coord([current_row, self.end_col_subjects], border=bc.get_border_thin()) \
                .draw(ws=self.ws, title=title_subject, font=text_mini, cell_len=Coord.CELL_LEN_AUTO_VERTICAL)
            self.end_col_subjects += 1
        self.end_col_subjects -= 1
        self.last_col = self.end_col_subjects

        Coord([3, 3], [3, self.end_col_subjects], border=bc.get_border_thin(top=bc.MEDIUM)) \
            .draw(ws=self.ws, title="Дисциплина", font=fc(bold=True), cell_len=Coord.CELL_LEN_AUTO_VERTICAL)

        last_row = current_row + num_students + 1
        self.last_row = last_row
        Coord([last_row, 1], [last_row, self.end_col_subjects], border=bc.get_border_medium()) \
            .draw(ws=self.ws, title="ВСЕГО", font=fc(bold=True), align=ac(horizontal=ac.RIGHT))

    def create_statistic(self):
        """Создание записи с успеваемостью"""
        print("Создание записи с успеваемостью")

        num_students = len(list(self.subjects.values())[0].keys())

        # Заполнение "Сред. балл"
        self.last_col += 1
        current_row = 3
        self.create_statistic_average(current_row, num_students)

        self.last_col += 1
        Coord([current_row, self.last_col], [current_row + 1, self.last_col],
              border=bc.get_border_medium(bottom=bc.THIN),
              font=fc(bold=True), title="Пропущено часов").draw(ws=self.ws)

        current_row += 2
        self.create_statistic_total(current_row, num_students)

    # --------------------------------------
    # Функции помощники create_statistic
    # --------------------------------------

    def create_statistic_average(self, start_row, num_rows):
        """Создание столбца [Сред.балл]"""
        print("Создание столбца [Сред.балл]")
        row = start_row
        Coord([row, self.last_col], [row + 1, self.last_col],
              border=bc.get_border_medium(bottom=bc.THIN), font=fc(bold=True), title="Сред. балл") \
            .draw(ws=self.ws, cell_len=10)

        row += 2
        letter_start = Coord.get_column_letter(self.start_col_subjects)
        letter_end = Coord.get_column_letter(self.end_col_subjects)

        for i in range(1, num_rows + 1):
            title = f"=AVERAGE({letter_start}{row}:{letter_end}{row})"
            Coord([row, self.last_col], border=bc.get_border_thin(left=bc.MEDIUM, right=bc.MEDIUM), font=fc(),
                  title=title).draw(ws=self.ws, type_size=Coord.CELL_SIZE_TYPE_MAX)
            row += 1

        letter = Coord.get_column_letter(self.last_col)
        title = f"=AVERAGE({letter}{start_row + 2}:{letter}{row - 1})"
        Coord([row, self.last_col], border=bc.get_border_medium(), font=fc(), title=title) \
            .draw(ws=self.ws, type_size=Coord.CELL_SIZE_TYPE_MAX, cell_len=10)

    def create_statistic_total(self, start_row, num_rows):
        """Создание столбца [Пропущено часов]"""
        print("Создание столбца [Пропущено часов]")
        row = start_row
        letter = Coord.get_column_letter(self.data['skip']['col_end'])
        for i in range(1, num_rows + 1):
            title = f"='{self.data['skip']['name_sheet']}'!{letter}{self.data['skip']['result_start_row'] + i - 1}"
            Coord([row, self.last_col], border=bc.get_border_thin(left=bc.MEDIUM, right=bc.MEDIUM), font=fc()) \
                .draw(ws=self.ws, type_size=Coord.CELL_SIZE_TYPE_MAX, title=title)
            row += 1

        letter = Coord.get_column_letter(self.last_col)
        title = f"=SUM({letter}{start_row + 1}:{letter}{row - 1})"
        Coord([row, self.last_col], border=bc.get_border_medium(), font=fc(), title=title) \
            .draw(ws=self.ws, type_size=Coord.CELL_SIZE_TYPE_MAX, cell_len=10)

    def create_title(self):
        """Создание титульной записи"""
        print("Создание титульной записи")
        Coord([1, 1], [1, self.last_col], ws=self.ws, title="Ведомость успеваемости и посещаемости", height=21,
              cell_len=5) \
            .draw(font=fc(size=fc.FONT_SIZE_BIG, bold=True))

        current_col = 1
        width_speciality = self.last_col - 10
        if width_speciality < 6:
            width_speciality = 5
        Coord([2, current_col], [2, width_speciality], ws=self.ws, title=self.speciality, height=30, cell_len=5) \
            .draw(font=fc(size=fc.FONT_SIZE_BIG, bold=True))
        current_col = width_speciality

        current_col += 1
        Coord([2, current_col], [2, current_col + 2], ws=self.ws, title="группа") \
            .draw(font=fc(size=fc.FONT_SIZE_BIG, bold=True), type_size=Coord.CELL_SIZE_TYPE_MAX)
        current_col += 2

        current_col += 1
        Coord([2, current_col], [2, current_col + 2], ws=self.ws, title=self.group) \
            .draw(font=fc(size=fc.FONT_SIZE_BIG, bold=True), type_size=Coord.CELL_SIZE_TYPE_MAX)
        current_col += 2

        current_col += 1
        Coord([2, current_col], [2, current_col + 3], ws=self.ws, title=self.month, cell_len=9) \
            .draw(font=fc(size=fc.FONT_SIZE_BIG, bold=True))
        current_col += 3

    def create_logical(self):
        """Создание столбцов отвечающих за вычисления"""

        num_students = len(list(self.subjects.values())[0].keys())

        self.last_col += 1
        current_row = 4
        self.create_logical_two(current_row, num_students)

        self.last_col += 1
        self.create_logical_two_three(current_row, num_students)

    def create_logical_two(self, start_row, num_rows):
        """Создание столбца [Учащиеся без "2"] """
        print('Создание столбца [Учащиеся без "2"]')
        row = start_row
        Coord([row, self.last_col], title='Учащиеся без 2', align=ac(horizontal=ac.LEFT)) \
            .draw(ws=self.ws)

        row += 1
        letter_start = Coord.get_column_letter(self.start_col_subjects)
        letter_end = Coord.get_column_letter(self.end_col_subjects)
        for i in range(1, num_rows + 1):
            title = f'=IF(COUNTIF({letter_start}{row}:{letter_end}{row},2),0,1)'
            Coord([row, self.last_col], font=fc(), title=title).draw(ws=self.ws, type_size=Coord.CELL_SIZE_TYPE_MAX)
            row += 1

        letter = Coord.get_column_letter(self.last_col)
        title = f"=SUM({letter}{start_row + 1}:{letter}{row - 1})"
        Coord([row, self.last_col], font=fc(), title=title).draw(ws=self.ws, cell_len=19)

    def create_logical_two_three(self, start_row, num_rows):
        """Создание столбца [Учащиеся без "2" и "3"]"""
        print('Создание столбца [Учащиеся без "2" и "3"]')
        row = start_row
        Coord([row, self.last_col], title='Учащиеся без 2 и 3', align=ac(horizontal=ac.LEFT)) \
            .draw(ws=self.ws)

        row += 1
        letter_start = Coord.get_column_letter(self.start_col_subjects)
        letter_end = Coord.get_column_letter(self.end_col_subjects)
        for i in range(1, num_rows + 1):
            title = f'=IF(COUNTIF({letter_start}{row}:{letter_end}{row},2) + COUNTIF({letter_start}{row}:{letter_end}{row},3),0,1)'
            Coord([row, self.last_col], font=fc(), title=title).draw(ws=self.ws, type_size=Coord.CELL_SIZE_TYPE_MAX)
            row += 1

        letter = Coord.get_column_letter(self.last_col)
        title = f"=SUM({letter}{start_row + 1}:{letter}{row - 1})"
        Coord([row, self.last_col], font=fc(), title=title).draw(ws=self.ws, cell_len=25)

    def create_result(self):
        """Создание столбца с итоговым результатом"""
        print("Создание столбца с итоговым результатом")

        current_row = self.last_row + 2
        coord_count_students = f'A{self.last_row - 1}'

        coord_not_two = f'{Coord.get_column_letter(self.last_col - 1)}{self.last_row}'
        title = f'=({coord_count_students}-{coord_not_two})/{coord_count_students}'
        Coord([current_row, 1], [current_row, 4], ws=self.ws, title="Успеваемость", cell_len=5) \
            .draw(align=ac(horizontal=ac.RIGHT))
        Coord([current_row, 5], ws=self.ws, title=title, height=16, cell_len=5, title_format=10) \
            .draw(align=ac(horizontal=ac.RIGHT), border=bc.get_border(bottom=bc.THIN),
                  type_size=Coord.CELL_SIZE_TYPE_MAX)
        current_row += 2

        coord_not_two_three = f'{Coord.get_column_letter(self.last_col)}{self.last_row}'
        title = f'={coord_not_two_three}/{coord_count_students}'
        Coord([current_row, 1], [current_row, 4], ws=self.ws, title="Качество", cell_len=5) \
            .draw(align=ac(horizontal=ac.RIGHT))
        Coord([current_row, 5], ws=self.ws, title=title, height=16, cell_len=5, title_format=10) \
            .draw(align=ac(horizontal=ac.RIGHT), border=bc.get_border(bottom=bc.THIN),
                  type_size=Coord.CELL_SIZE_TYPE_MAX)

        current_row += 2

        coord_total_hours = f'{Coord.get_column_letter(self.end_col_subjects + 2)}{self.last_row}'
        title = f'={coord_total_hours}/{coord_count_students}'
        Coord([current_row, 1], [current_row, 4], ws=self.ws, title="На одного студента (кол-во пропусков)", cell_len=5) \
            .draw(align=ac(horizontal=ac.RIGHT))
        Coord([current_row, 5], ws=self.ws, title=title, height=16, cell_len=5, title_format=2) \
            .draw(align=ac(horizontal=ac.RIGHT), border=bc.get_border(bottom=bc.THIN),
                  type_size=Coord.CELL_SIZE_TYPE_MAX)
