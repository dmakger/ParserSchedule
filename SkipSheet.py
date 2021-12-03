from openpyxl import Workbook

from SheetHelper import SheetHelper as sh
from BorderCoord import BorderCoord as bc
from FontCoord import FontCoord as fc
from AlignmentCoord import AlignmentCoord as ac
from Coord import Coord


class SkipSheet:

    def __init__(self, wb: Workbook, subjects: dict, schedule: dict, month: int, year: int, all_days: list,
                 name_sheet: str = None):
        if name_sheet is None:
            name_sheet = "(ЭН)"
        self.name_sheet = name_sheet
        self.wb = wb
        self.subjects = subjects
        self.schedule = schedule
        self.month = month
        self.year = year
        self.all_days = all_days

        self.ws = None
        self.col_end = None
        self.data_schedule = None
        self.last_row = 1
        self.result_start_row = 4

        self.create_sheet()

    def create_sheet(self):
        """Создание листа"""
        self.ws = self.wb.create_sheet(self.name_sheet)

        self.create_students(start_row=1, start_col=1)
        self.create_skip(start_row=1, start_col=3)

    def create_students(self, start_row, start_col):
        """Создание записи со студентами"""
        print("Создание записи со студентами")
        row = start_row + 2
        col = start_col
        # Title
        Coord([start_row, col], [row, col], border=bc.get_border_thin(top=bc.BOLD, left=bc.BOLD), font=fc(bold=True)) \
            .draw(ws=self.ws, title="№ п/п")
        col += 1

        Coord([start_row, col], [row, col], border=bc.get_border_thin(top=bc.BOLD), font=fc(bold=True)) \
            .draw(ws=self.ws, title="ФИО")

        # Main
        row += 1
        students = list(list(self.subjects.values())[0].keys())
        for i in range(len(students)):
            Coord([row, start_col], border=bc.get_border_thin(left=bc.BOLD, bottom=bc.NONE)).draw(ws=self.ws,
                                                                                                  title=str(i + 1))
            Coord([row, col], border=bc.get_border_thin(bottom=bc.NONE)) \
                .draw(ws=self.ws, title=students[i], align=ac(horizontal=ac.LEFT), type_size=Coord.CELL_SIZE_TYPE_MAX)
            row += 1

        self.last_row = row

        Coord([self.last_row, start_col], [self.last_row, start_col + 1], title="Наименование Предмета",
              border=bc.get_border_bold(right=bc.THIN), ws=self.ws, font=fc(bold=True, size=fc.FONT_SIZE_MINI),
              cell_len=4, height=50).draw()

    def create_skip(self, start_row, start_col):
        """Создание записи с парами и пропусками"""
        row = start_row
        col = start_col
        self.create_skip_title(row, col)

        row = start_row + 3
        self.create_skip_skip(row, col)

    def create_skip_title(self, start_row, start_col):
        """Создание титульной записи с датами и предметами"""
        print("Создание титульной записи с датами и предметами")
        row = start_row
        col = start_col

        count_days = 0
        self.data_schedule = dict()
        for week in self.schedule:
            for day_week in range(1, 8):
                day = week.get(day_week, -1)
                if day != -1:
                    update_day = sh.get_lessons(day)

                    border_empty = bc.get_border(top=bc.BOLD, bottom=bc.MEDIUM)
                    border_number_lesson = bc.get_border_thin()
                    border_subjects = bc.get_border_bold(left=bc.NONE, right=bc.NONE)

                    lessons = sorted(update_day.keys())
                    last_lesson = lessons[-1]
                    for lesson in lessons:
                        if lesson == last_lesson:
                            border_empty = bc.get_border_medium(top=bc.BOLD, left=bc.NONE)
                            border_number_lesson = bc.get_border_thin(right=bc.MEDIUM)
                            border_subjects = bc.get_border_bold(left=bc.NONE, right=bc.MEDIUM)
                        # Пустые колонки сверху
                        Coord([row, col], ws=self.ws, border=border_empty).draw(
                            font=fc(bold=True, size=fc.FONT_SIZE_MINI), cell_len=Coord.CELL_LEN_AUTO_VERTICAL)
                        # Номер пары
                        Coord([row + 2, col], ws=self.ws, title=sh.get_char_lesson(lesson), font=fc(bold=True)) \
                            .draw(cell_len=8, border=border_number_lesson)
                        # Название пары
                        Coord([self.last_row, col], ws=self.ws, title=update_day[lesson], border=border_subjects) \
                            .draw(font=fc(bold=True, size=fc.FONT_SIZE_MINI), cell_len=Coord.CELL_LEN_AUTO_VERTICAL)
                        col += 1
                    # Дата
                    Coord([row + 1, col - len(lessons)], [row + 1, col - 1], ws=self.ws, font=fc(bold=True), cell_len=8,
                          border=bc.get_border(right=bc.MEDIUM), title=self.all_days[count_days]).draw()
                    self.data_schedule[self.all_days[count_days]] = col - 1
                    count_days += 1

        Coord([start_row, col], [start_row + 2, col], title="ПРОПУЩЕНО (часов)", ws=self.ws,
              font=fc(bold=True, size=fc.FONT_SIZE_MINI), type_size=Coord.CELL_LEN_AUTO_VERTICAL,
              border=bc.get_border_bold(left=bc.MEDIUM, bottom=bc.THIN)).draw()
        self.col_end = col - 1

    def create_skip_skip(self, start_row, start_col):
        """Создание записи с пропусками"""
        print("Создание записи с пропусками")
        row = start_row
        col = start_col
        row_lesson = self.last_row
        row_date = start_row - 2

        name_students: list = sorted(list(list(self.subjects.values())[0].keys()))
        last_row = len(name_students) + start_row
        for i in range(start_row, last_row):
            for j in range(start_col, self.col_end + 1):
                Coord([i, j], ws=self.ws, border=bc.get_border_thin()).draw()

        last_date = None
        result_dict = dict()
        for title_lesson in self.subjects:
            print(f'Формирование записи с предметом "{title_lesson}"')
            for name, student_subjects in self.subjects[title_lesson].items():
                row_student = name_students.index(name) + start_row
                for date, result in student_subjects.items():
                    for col_day in range(col, self.col_end + 1):
                        current_date = Coord.get_title(ws=self.ws, row=row_date, column=col_day)
                        if current_date is None:
                            current_date = last_date
                        else:
                            last_date = current_date
                        if self.data_schedule[current_date] == col_day:
                            Coord([row_student, col_day], ws=self.ws, border=bc.get_border_thin(right=bc.MEDIUM)) \
                                .draw(type_size=Coord.CELL_SIZE_TYPE_MAX)
                        if date == current_date and \
                                title_lesson == Coord.get_title(ws=self.ws, row=row_lesson, column=col_day) and \
                                result == "Н":
                            result_dict[name] = result_dict.get(name, 0) + 1
                            Coord([row_student, col_day], ws=self.ws, title=result, border=bc.get_border_thin()) \
                                .draw(type_size=Coord.CELL_SIZE_TYPE_MAX)

        self.col_end += 1
        letter_start = Coord.get_column_letter(start_col)
        letter_end = Coord.get_column_letter(self.col_end - 1)
        for row in range(start_row, last_row):
            title = f'=(COUNTIF({letter_start}{row}:{letter_end}{row},"н"))*2'
            Coord([row, self.col_end], ws=self.ws, title=title, cell_len=13, height=15.8)\
                .draw(border=bc.get_border_thin(left=bc.BOLD, right=bc.BOLD))

        letter = Coord.get_column_letter(self.col_end)
        title = f'=SUM({letter}{start_row}:{letter}{last_row - 1})'
        Coord([last_row, self.col_end], ws=self.ws, title=title, cell_len=13)\
            .draw(border=bc.get_border_bold(top=bc.THIN))
