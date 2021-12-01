from typing import Union

from openpyxl.worksheet._write_only import WriteOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

from BorderCoord import BorderCoord as bc
from FontCoord import FontCoord as fc
from AlignmentCoord import AlignmentCoord as ac
from openpyxl.styles import Border


class Coord:
    CELL_LEN_NONE = 0
    CELL_LEN_AUTO_VERTICAL = 1
    CELL_LEN_AUTO_HORIZONTAL = 2

    CELL_WIDTH = 8.43
    CELL_HEIGHT = 15.75

    CELL_SIZE_TYPE_NONE = 0
    CELL_SIZE_TYPE_MAX = 1

    @staticmethod
    def get_column_letter(column: int):
        return get_column_letter(column)

    @staticmethod
    def get_title(ws, row, column):
        return ws.cell(row=row, column=column).value

    def __init__(self, start: list, end: list = None, font: fc = None, align: ac = None,
                 border: Border = None, width: float = CELL_WIDTH, height: float = CELL_HEIGHT, title=None,
                 cell_len: int = CELL_LEN_AUTO_HORIZONTAL, type_size: int = CELL_SIZE_TYPE_NONE,
                 ws: Union[WriteOnlyWorksheet, Worksheet] = None, title_format: int = None):
        self.start = start
        self.start_row: int = start[0]
        self.start_col: int = start[1]

        if end is None:
            end = start
        self.end = end
        self.end_row: int = end[0]
        self.end_col: int = end[1]

        if font is None:
            font = fc()
        self.font = font

        if align is None:
            align = ac()
        self.align = align

        if ws is None:
            ws = None
        self.ws = ws

        if border is None:
            border = bc.get_border()
        elif type(border) == dict:
            border = bc.get_border(left=border.get('left', bc.NONE),
                                   right=border.get('right', bc.NONE),
                                   top=border.get('top', bc.NONE),
                                   bottom=border.get('bottom', bc.NONE))
        self.border = border

        self.title = title
        self.title_format = title_format
        self.cell_len = cell_len
        self.width = width
        self.height = height
        self.type_size = type_size

    def draw(self, ws: Union[WriteOnlyWorksheet, Worksheet] = None, title: str = None, font: fc = None,
             align: ac = None, type_size: int = None, cell_len: float = None, border: Border = None):
        """Отрисовка текста"""
        if ws is None:
            ws = self.ws
        if title is None:
            title = self.title
        if font is None:
            font = self.font
        if align is None:
            align = self.align
        if type_size is None:
            type_size = self.type_size
        if cell_len is None:
            cell_len = self.cell_len
        if border is None:
            border = self.border
        elif type(border) == dict:
            border = bc.get_border(left=border.get('left', bc.NONE),
                                   right=border.get('right', bc.NONE),
                                   top=border.get('top', bc.NONE),
                                   bottom=border.get('bottom', bc.NONE))

        ws.merge_cells(
            start_row=self.start_row, start_column=self.start_col,
            end_row=self.end_row, end_column=self.end_col
        )
        cell = ws.cell(row=self.start_row, column=self.start_col)
        cell.font = font.get()
        cell.alignment = align.get()
        self.set_border(ws, border)

        self.set_cell_size(ws=ws, title=title, cell=cell, cell_len=cell_len, type_size=type_size)
        self.set_title(ws=ws, title=title, cell=cell, title_format=self.title_format)

    def set_border(self, ws: Union[WriteOnlyWorksheet, Worksheet] = None, border: Border = None):
        if ws is None:
            ws = self.ws
        if border is None:
            border = self.border
        elif type(border) == dict:
            border = bc.get_border(left=border.get('left', bc.NONE),
                                   right=border.get('right', bc.NONE),
                                   top=border.get('top', bc.NONE),
                                   bottom=border.get('bottom', bc.NONE))

        for i in range(self.start_row, self.end_row + 1):
            for j in range(self.start_col, self.end_col + 1):
                ws.cell(row=i, column=j).border = border

    def get_size_column(self, title, cell_len=CELL_LEN_AUTO_HORIZONTAL):
        if title is None:
            title = self.title
        title = str(title)

        # return max([len(word) for word in title.split(' ')]) + 10
        if cell_len == self.CELL_LEN_NONE:
            return self.CELL_WIDTH
        elif cell_len == self.CELL_LEN_AUTO_HORIZONTAL:
            cell_len_digit = (len(title) + 2) * 1.2
            # if self.start_row == self.end_row:
            return cell_len_digit

        elif cell_len == self.CELL_LEN_AUTO_VERTICAL:
            return (max([len(word) for word in title.split(' ')]) + 2) * 1.2
        else:
            return cell_len

    def set_cell_size(self, cell, cell_len, ws=None, title=None, width_column: int = None,
                      type_size: int = CELL_SIZE_TYPE_NONE):
        if title is None:
            title = self.title
        if ws is None:
            ws: Union[WriteOnlyWorksheet, Worksheet] = self.ws
        if width_column is None:
            width_column = self.get_size_column(title, cell_len)

        height = self.get_size_row(ws=ws, title=title, width_column=width_column)
        # print(title, height)

        # Устанавливаем дефолтные значения
        if ws.column_dimensions[cell.column_letter].width is None:
            ws.column_dimensions[cell.column_letter].width = self.CELL_WIDTH
        if ws.row_dimensions[cell.row].height is None:
            ws.row_dimensions[cell.row].height = self.CELL_HEIGHT

        if type_size == self.CELL_SIZE_TYPE_MAX:
            if ws.column_dimensions[cell.column_letter].width < width_column:
                ws.column_dimensions[cell.column_letter].width = width_column
            else:
                width_column = ws.column_dimensions[cell.column_letter].width
            self.width = width_column

            if ws.row_dimensions[cell.row].height > height:
                ws.row_dimensions[cell.row].height = height
            else:
                height = ws.row_dimensions[cell.row].height
            self.height = height

        elif type_size == self.CELL_SIZE_TYPE_NONE:
            ws.column_dimensions[cell.column_letter].width = width_column
            ws.row_dimensions[cell.row].height = height

    def get_size_row(self, width_column: int, title: str = None, ws: Union[WriteOnlyWorksheet, Worksheet] = None):
        if title is None:
            title = self.title
        title = str(title)
        if ws is None:
            ws: Union[WriteOnlyWorksheet, Worksheet] = self.ws

        if self.height != self.CELL_HEIGHT:
            return self.height

        current_row = ws.row_dimensions[self.start_row].height
        if current_row is None:
            current_row = self.CELL_HEIGHT

        height = int(round(len(title) / width_column) * self.CELL_HEIGHT) + 10
        if current_row + 10 > height:
            return current_row
        return height

    def set_title(self, cell, ws: Union[WriteOnlyWorksheet, Worksheet] = None, title: str = None,
                  title_format: int = None):
        if ws is None:
            ws = self.ws
        if title is None:
            title = self.title
        if not (title_format is None):
            cell.number_format = numbers.BUILTIN_FORMATS[title_format]

        type_title = type(title)
        if type_title == int:
            cell.value = title
        elif type_title == str:
            if len(title) > 0 and title[0] == '=':
                ws[cell.coordinate] = title
            else:
                cell.value = title
