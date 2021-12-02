from openpyxl import Workbook

from SheetHelper import SheetHelper as sh
from BorderCoord import BorderCoord as bc
from FontCoord import FontCoord as fc
from AlignmentCoord import AlignmentCoord as ac
from Coord import Coord


class SubjectsSheet:

    def __init__(self, wb: Workbook, subjects: dict, schedule: dict, month: int, year: int, all_days: list,
                 name_sheet: str = None):
        if name_sheet is None:
            name_sheet = "(БН)"
        self.name_sheet = name_sheet
        self.wb = wb
        self.subjects = subjects
        self.schedule = schedule
        self.month = month
        self.year = year
        self.all_days = all_days

        self.ws = None
        self.col_end = None
        self.data_schedule = None
        self.last_row = 1

        self.create_sheet()

    def create_sheet(self):
        """Создание листа"""
        self.ws = self.wb.create_sheet(self.name_sheet)

        self.create_students(start_row=1, start_col=1)
        self.create_subjects(start_row=1, start_col=3)
        self.create_last_rows(start_row=self.last_row, start_col=1)

    def create_students(self, start_row, start_col):
        """Создание записи со студентами"""
        row = start_row + 2
        col = start_col
        # Title
        Coord([start_row, col], [row, col], border=bc.get_border_thin(top=bc.BOLD, left=bc.BOLD), font=fc(bold=True)) \
            .draw(ws=self.ws, title="№ п/п")
        col += 1

        Coord([start_row, col], [row, col], border=bc.get_border_thin(top=bc.BOLD), font=fc(bold=True)) \
            .draw(ws=self.ws, title="ФИО")

        # Main
        row += 1
        students = list(list(self.subjects.values())[0].keys())
        for i in range(len(students)):
            Coord([row, start_col], border=bc.get_border_thin(left=bc.BOLD, bottom=bc.NONE)).draw(ws=self.ws,
                                                                                                  title=str(i + 1))
            Coord([row, col], border=bc.get_border_thin(bottom=bc.NONE)) \
                .draw(ws=self.ws, title=students[i], align=ac(horizontal=ac.LEFT), type_size=Coord.CELL_SIZE_TYPE_MAX)
            row += 1
        # Coord([row, start_col], ws=self.ws, border=bc.get_border(top=bc.BOLD),
        #       type_size=Coord.CELL_SIZE_TYPE_MAX).draw()
        # Coord([row, col], ws=self.ws, border=bc.get_border(top=bc.BOLD), type_size=Coord.CELL_SIZE_TYPE_MAX).draw()

        self.last_row = row

    def create_subjects(self, start_row, start_col):
        """Создание записи с парами и оценками"""

        row = start_row
        col = start_col
        self.create_subjects_title(row, col)

        row = start_row + 3
        # self.create_skip_skip(row, col)

    def create_subjects_title(self, start_row, start_col):
        """Создание титульной записи с датами и предметами"""
        row = start_row
        col = start_col

        count_days = 0
        self.data_schedule = dict()
        for week in self.schedule:
            for day_week in range(1, 8):
                day = week.get(day_week, -1)
                if day != -1:
                    update_day = sh.get_lessons(day)

                    border_subjects = bc.get_border(top=bc.BOLD, bottom=bc.MEDIUM)
                    border_number_lesson = bc.get_border_thin()
                    border_signature = bc.get_border(top=bc.THIN, bottom=bc.BOLD)

                    lessons = sorted(update_day.keys())
                    last_lesson = lessons[-1]
                    for lesson in lessons:
                        if lesson == last_lesson:
                            border_subjects = bc.get_border_medium(top=bc.BOLD, left=bc.NONE)
                            border_number_lesson = bc.get_border_thin(right=bc.MEDIUM)
                            border_signature = bc.get_border_medium(bottom=bc.BOLD, top=bc.THIN, left=bc.NONE)
                        # Пустые колонки сверху
                        Coord([row, col], ws=self.ws, border=border_subjects).draw(
                            font=fc(bold=True, size=fc.FONT_SIZE_MINI), cell_len=Coord.CELL_LEN_AUTO_VERTICAL)
                        # Номер пары
                        Coord([row + 2, col], ws=self.ws, title=sh.get_char_lesson(lesson), font=fc(bold=True)) \
                            .draw(cell_len=8, border=border_number_lesson)
                        # Название пары
                        Coord([self.last_row, col], ws=self.ws, title=update_day[lesson], border=border_subjects) \
                            .draw(font=fc(bold=True, size=fc.FONT_SIZE_MINI), cell_len=Coord.CELL_LEN_AUTO_VERTICAL)
                        # Для подписи
                        Coord([self.last_row + 1, col], ws=self.ws, border=border_signature) \
                            .draw(font=fc(bold=True, size=fc.FONT_SIZE_MINI), cell_len=Coord.CELL_LEN_AUTO_VERTICAL)
                        col += 1
                    # Дата
                    Coord([row + 1, col - len(lessons)], [row + 1, col - 1], ws=self.ws, font=fc(bold=True), cell_len=8,
                          border=bc.get_border(right=bc.MEDIUM), title=self.all_days[count_days]).draw()
                    self.data_schedule[self.all_days[count_days]] = col - 1
                    count_days += 1

        self.col_end = col - 1

    def create_last_rows(self, start_row, start_col):
        row = start_row + 1
        col = start_col + 1
        Coord([start_row, start_col], [start_row, col], title="Наименование Предмета",
              border=bc.get_border_thin(top=bc.BOLD, left=bc.BOLD), ws=self.ws,
              font=fc(bold=True, size=fc.FONT_SIZE_MINI), cell_len=4, height=50).draw()
        Coord([row, start_col], [row, col], title="Подпись преподавателя",
              border=bc.get_border_thin(bottom=bc.BOLD, left=bc.BOLD), ws=self.ws,
              font=fc(bold=True, size=fc.FONT_SIZE_MINI), cell_len=4, height=50).draw()
        self.last_row = row